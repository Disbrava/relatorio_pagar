import datetime
import os
from uuid import UUID, uuid4

import pandas as pd
from reportlab.platypus import SimpleDocTemplate
from sqlalchemy.orm import Session
from typing import List, Callable, Dict, Union

from app.database.sync_database import SyncDatabase
from app.domain.services.data_extractor_interface import DataExtractorServiceInterface
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.pagesizes import A4, landscape
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle, SimpleDocTemplate, Spacer, PageBreak
from reportlab.lib.units import inch
from reportlab.platypus import Table, TableStyle, Paragraph


class OracleDataExtractorServiceService(DataExtractorServiceInterface):

    def __init__(self, database: SyncDatabase):
        self._database = database

    def __transform_data(self, data: pd.DataFrame, file_type: str) -> Union[
        List[pd.DataFrame], Dict[str, pd.DataFrame]]:
        columns = ['LOJA', 'NOME', 'CATEGORIA', 'EMISSAO', 'VENCIMENTO', 'DEPARTAMENTO', 'INFORMACAO',
                   'VALOR DO TÍTULO',
                   'SALDO']
        data.columns = columns
        data = data[['LOJA', 'NOME', 'CATEGORIA', 'EMISSAO', 'SALDO']]

        maraba = data[data['LOJA'].isin(['1.1', '1.2'])]
        total_balance = pd.DataFrame([[maraba['SALDO'].sum()]], columns=['SALDO'])
        maraba = pd.concat([maraba, total_balance])

        nacional = data[data['LOJA'].isin(['2.1', '2.2', '2.4'])]
        total_balance = pd.DataFrame([[nacional['SALDO'].sum()]], columns=['SALDO'])
        nacional = pd.concat([nacional, total_balance])

        byd = data[data['LOJA'].isin(['40.1'])]
        total_balance = pd.DataFrame([[byd['SALDO'].sum()]], columns=['SALDO'])
        byd = pd.concat([byd, total_balance])

        if file_type == 'xlsx':
            return [maraba, nacional, byd]

        return {'MARABÁ': maraba, 'NACIONAL': nacional, 'BYD': byd}

    def __transform_data_in_xlsx_file(self, data: List[pd.DataFrame]) -> None:
        attach_name = f'Relatorio_Pagar_{str(datetime.datetime.now().day) + "_" + str(datetime.datetime.now().month) + "_" + str(datetime.datetime.now().year)}.xlsx'
        if not os.path.exists('tmp'):
            os.makedirs('tmp')

            # Agora você pode salvar o arquivo
        file_path = os.path.join('tmp', attach_name)
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            data[0].to_excel(writer, index=False, sheet_name='MARABÁ')
            data[1].to_excel(writer, index=False, sheet_name='NACIONAL')
            data[2].to_excel(writer, index=False, sheet_name='BYD')
            workbook = writer.book
            worksheet_maraba = writer.sheets['MARABÁ']
            worksheet_nacional = writer.sheets['NACIONAL']
            worksheet_byd = writer.sheets['BYD']

            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")

            for cell in worksheet_maraba[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            for cell in worksheet_nacional[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            for cell in worksheet_byd[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            thin_border = Border(left=Side(style='thick'),
                                 right=Side(style='thick'),
                                 top=Side(style='thick'),
                                 bottom=Side(style='thick'))
            brazilian_coin_format = 'R$ #,##0.00'

            for sheet in [worksheet_maraba, worksheet_nacional, worksheet_byd]:
                for cell in sheet['H']:
                    if cell.row > 1:
                        cell.number_format = brazilian_coin_format

                for cell in sheet['I']:
                    if cell.row > 1:
                        cell.number_format = brazilian_coin_format

            bold_font = Font(bold=True)
            fill_color = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            for sheet in [worksheet_maraba, worksheet_nacional, worksheet_byd]:
                last_column = sheet.max_column
                last_row = sheet.max_row
                last_cell = sheet.cell(row=last_row, column=last_column)
                last_cell.font = bold_font
                last_cell.fill = fill_color

    def _transform_data_in_pdf_file(self, data: Dict[str, pd.DataFrame]) -> UUID:

        attach_name = f'Relatorio_Pagar_{str(datetime.datetime.now().day) + "_" + str(datetime.datetime.now().month) + "_" + str(datetime.datetime.now().year)}.pdf'
        if not os.path.exists('tmp'):
            os.makedirs('tmp')

        # Arquivo
        file_path = os.path.join('tmp', attach_name)
        pdf = SimpleDocTemplate(file_path, pagesize=landscape(A4))
        title = "Relatório de títulos a Pagar na Data  " + datetime.datetime.now().strftime("%d-%m-%Y")
        file_id = uuid4()
        elements = []
        styles = getSampleStyleSheet()
        style_normal = styles["BodyText"]
        elements.append(Paragraph(f"{title.upper()}", styles['Title']))
        elements.append(Spacer(1, 20))

        for key, table in data.items():
            table = table.fillna('')
            table['SALDO'] = table['SALDO'].map('R$ {:,.2f}'.format)

            table_data = [table.columns.tolist()] + [[Paragraph(str(cell), styles['Normal']) for cell in row] for row in
                                                     table.values.tolist()]
            num_columns = len(table.columns)
            max_table_width = landscape(A4)[0] - 2 * inch  # Largura máxima da tabela com margens
            col_widths = [max_table_width / num_columns] * num_columns

            table = Table(table_data, colWidths=col_widths)
            last_row = len(table_data) - 1

            table_style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),  # Fundo do cabeçalho
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),  # Cor do texto do cabeçalho
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),  # Alinhamento central
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),  # Fonte do cabeçalho
                ('FONTSIZE', (0, 0), (-1, -1), 9),  # Tamanho da fonte
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),  # Espaçamento inferior no cabeçalho
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),  # Cor de fundo das linhas
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),  # Linhas de grade
                ('BACKGROUND', (num_columns - 1, last_row), (num_columns - 1, last_row), colors.yellow),
                ('ALIGN', (num_columns - 1, last_row), (num_columns - 1, last_row), 'CENTER'),  # Alinhamento central
                ('FONTNAME', (num_columns - 1, last_row), (num_columns - 1, last_row), 'Helvetica-Bold'),
                ('FONTSIZE', (num_columns - 1, last_row), (num_columns - 1, last_row), 12),
                ('TEXTCOLOR', (num_columns - 1, last_row), (num_columns - 1, last_row), colors.black),
            ])
            table.setStyle(table_style)
            table.splitByRow = True  # Permite dividir a tabela em múltiplas páginas
            table.repeatRows = 1  # Número de linhas a serem repetidas (cabeçalho)

            subtitle = Paragraph(f"{str(key)}", styles['Heading2'])
            elements.append(subtitle)  # Adiciona o título
            elements.append(Spacer(1, 5))
            elements.append(table)  # Adiciona a tabela
            elements.append(Spacer(1, 20))  # Espaçamento entre tabelas
            elements.append(PageBreak())

        pdf.build(elements)

        return file_id

    def extract_data_to_report(self, data: Dict[str, pd.DataFrame]) -> Union[UUID, str]:
        try:
            result = self._transform_data_in_pdf_file(data)
            return result
        except Exception as e:
            return f"Error: {e.__str__()}"

