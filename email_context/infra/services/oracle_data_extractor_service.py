import datetime
import os

import pandas as pd
from reportlab.platypus import SimpleDocTemplate
from sqlalchemy.orm import Session
from typing import List, Callable, Dict, Union

from email_context.database.sync_database import SyncDatabase
from email_context.domain.services.data_extractor_interface import DataExtractorInterface
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.pagesizes import A4, landscape
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle, SimpleDocTemplate, Spacer, PageBreak
from reportlab.lib.units import inch
from reportlab.platypus import Table, TableStyle, Paragraph


class OracleDataExtractorService(DataExtractorInterface):

    def __init__(self, database: SyncDatabase):
        self._database = database

    def __transform_data(self, data: pd.DataFrame, file_type: str) -> Union[List[pd.DataFrame], Dict[str, pd.DataFrame]]:
        columns = ['LOJA', 'NOME', 'CATEGORIA', 'EMISSAO', 'VENCIMENTO', 'DEPARTAMENTO', 'INFORMACAO',
                   'VALOR DO TÍTULO',
                   'SALDO']
        data.columns = columns
        data = data[['LOJA', 'NOME', 'CATEGORIA', 'EMISSAO',  'SALDO']]

        maraba = data[data['LOJA'].isin(['1.1', '1.2'])]
        total_balance = pd.DataFrame([[maraba['SALDO'].sum()]], columns=['SALDO'])
        maraba = pd.concat([maraba, total_balance])

        nacional = data[data['LOJA'].isin(['2.1', '2.2', '2.4'])]
        total_balance = pd.DataFrame([[nacional['SALDO'].sum()]], columns=['SALDO'])
        nacional = pd.concat([nacional, total_balance])

        byd = data[data['LOJA'].isin(['40.1'])]
        total_balance = pd.DataFrame([[byd['SALDO'].sum()]], columns=['SALDO'])
        byd = pd.concat([byd, total_balance])

        if file_type == 'xlsx':
            return [maraba, nacional, byd]

        return {'MARABÁ': maraba, 'NACIONAL': nacional, 'BYD': byd}

    def __transform_data_in_xlsx_file(self, data: List[pd.DataFrame]) -> None:
        attach_name = f'Relatorio_Pagar_{str(datetime.datetime.now().day) + "_" + str(datetime.datetime.now().month) + "_" + str(datetime.datetime.now().year)}.xlsx'
        if not os.path.exists('tmp'):
            os.makedirs('tmp')

            # Agora você pode salvar o arquivo
        file_path = os.path.join('tmp', attach_name)
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            data[0].to_excel(writer, index=False, sheet_name='MARABÁ')
            data[1].to_excel(writer, index=False, sheet_name='NACIONAL')
            data[2].to_excel(writer, index=False, sheet_name='BYD')
            workbook = writer.book
            worksheet_maraba = writer.sheets['MARABÁ']
            worksheet_nacional = writer.sheets['NACIONAL']
            worksheet_byd = writer.sheets['BYD']

            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")

            for cell in worksheet_maraba[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            for cell in worksheet_nacional[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            for cell in worksheet_byd[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            thin_border = Border(left=Side(style='thick'),
                                 right=Side(style='thick'),
                                 top=Side(style='thick'),
                                 bottom=Side(style='thick'))
            brazilian_coin_format = 'R$ #,##0.00'

            for sheet in [worksheet_maraba, worksheet_nacional, worksheet_byd]:
                for cell in sheet['H']:
                    if cell.row > 1:
                        cell.number_format = brazilian_coin_format

                for cell in sheet['I']:
                    if cell.row > 1:
                        cell.number_format = brazilian_coin_format

            bold_font = Font(bold=True)
            fill_color = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            for sheet in [worksheet_maraba, worksheet_nacional, worksheet_byd]:
                last_column = sheet.max_column
                last_row = sheet.max_row
                last_cell = sheet.cell(row=last_row, column=last_column)
                last_cell.font = bold_font
                last_cell.fill = fill_color

    def __transform_data_in_pdf_file(self, data: Dict[str, pd.DataFrame]) -> None:

        attach_name = f'Relatorio_Pagar_{str(datetime.datetime.now().day) + "_" + str(datetime.datetime.now().month) + "_" + str(datetime.datetime.now().year)}.pdf'
        if not os.path.exists('tmp'):
            os.makedirs('tmp')

        # Arquivo
        file_path = os.path.join('tmp', attach_name)
        pdf = SimpleDocTemplate(file_path, pagesize=landscape(A4))
        title = "Relatório de títulos a Pagar na Data  " + datetime.datetime.now().strftime("%d-%m-%Y")

        elements = []
        styles = getSampleStyleSheet()
        style_normal = styles["BodyText"]
        elements.append(Paragraph(f"{title.upper()}", styles['Title']))
        elements.append(Spacer(1, 20))

        for key, table in data.items():
            table = table.fillna('')
            table['SALDO'] = table['SALDO'].map('R$ {:,.2f}'.format)

            table_data = [table.columns.tolist()] + [[Paragraph(str(cell), styles['Normal']) for cell in row] for row in
                                                     table.values.tolist()]
            num_columns = len(table.columns)
            max_table_width = landscape(A4)[0] - 2 * inch  # Largura máxima da tabela com margens
            col_widths = [max_table_width / num_columns] * num_columns

            table = Table(table_data, colWidths=col_widths)
            last_row = len(table_data) - 1

            table_style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),  # Fundo do cabeçalho
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),  # Cor do texto do cabeçalho
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),  # Alinhamento central
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),  # Fonte do cabeçalho
                ('FONTSIZE', (0, 0), (-1, -1), 9),  # Tamanho da fonte
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),  # Espaçamento inferior no cabeçalho
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),  # Cor de fundo das linhas
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),  # Linhas de grade
                ('BACKGROUND', (num_columns - 1, last_row), (num_columns - 1, last_row), colors.yellow),
                ('ALIGN', (num_columns - 1, last_row), (num_columns - 1, last_row), 'CENTER'),  # Alinhamento central
                ('FONTNAME', (num_columns - 1, last_row), (num_columns - 1, last_row), 'Helvetica-Bold'),
                ('FONTSIZE', (num_columns - 1, last_row), (num_columns - 1, last_row), 12),
                ('TEXTCOLOR', (num_columns - 1, last_row), (num_columns - 1, last_row), colors.black),
            ])
            table.setStyle(table_style)
            table.splitByRow = True  # Permite dividir a tabela em múltiplas páginas
            table.repeatRows = 1  # Número de linhas a serem repetidas (cabeçalho)

            subtitle = Paragraph(f"{str(key)}", styles['Heading2'])
            elements.append(subtitle)  # Adiciona o título
            elements.append(Spacer(1, 5))
            elements.append(table)  # Adiciona a tabela
            elements.append(Spacer(1, 20))  # Espaçamento entre tabelas
            elements.append(PageBreak())

        pdf.build(elements)

    def extract_data_to_report(self) -> List[pd.DataFrame]:
        sql_query = '''
        WITH dw AS (
        SELECT DISTINCT
                	A.CLIENTE,
                	A.NOME,
                	B.EMPRESA,
                	B.REVENDA,
                	B.TITULO,
                	B.VAL_TITULO,
                	B.DTA_EMISSAO,
                	B.DTA_VENCIMENTO,
                	B.STATUS,
                	B.DEPARTAMENTO,
                	B.ORIGEM,
                	B.TIPO,
                	(
                        SELECT
                            FO.DES_ORIGEM
                        FROM
                            CNP.FIN_ORIGEM FO
                        WHERE
                            FO.ORIGEM = B.ORIGEM
                            AND FO.EMPRESA = B.EMPRESA
                            AND FO.REVENDA = B.REVENDA
                        GROUP BY
                            FO.DES_ORIGEM
                    ) DES_ORIGEM,
                    B.PAGAR_RECEBER,
                    B.BANCO,
                    (
                        SELECT DES_BANCO FROM CNP.FIN_BANCO 
                        WHERE CNP.FIN_BANCO.BANCO=B.BANCO 
                        AND CNP.FIN_BANCO.EMPRESA = B.EMPRESA 
                        AND CNP.FIN_BANCO.REVENDA = B.REVENDA
                    ) DES_BANCO,
                    CASE
                        WHEN B.STATUS = 'CA' THEN 0
                        ELSE (
                            B.VAL_TITULO - (
                                COALESCE(
                                    (
                                        SELECT
                                            SUM(CNP.FIN_TITULO_PAGAMENTO.VAL_PAGAMENTO)
                                        FROM
                                            CNP.FIN_TITULO_PAGAMENTO
                                        WHERE
                                            CNP.FIN_TITULO_PAGAMENTO.EMPRESA = B.EMPRESA
                                            AND CNP.FIN_TITULO_PAGAMENTO.REVENDA = B.REVENDA
                                            AND CNP.FIN_TITULO_PAGAMENTO.TITULO = B.TITULO
                                            AND CNP.FIN_TITULO_PAGAMENTO.DUPLICATA = B.DUPLICATA
                                            AND CNP.FIN_TITULO_PAGAMENTO.CLIENTE = B.CLIENTE
                                            AND CNP.FIN_TITULO_PAGAMENTO.TIPO = B.TIPO
                                    ),
                                    0
                                ) + COALESCE(
                                    (
                                        SELECT
                                            SUM(CNP.FIN_DEVOLUCAO.VAL_DEVOLUCAO)
                                        FROM
                                            CNP.FIN_DEVOLUCAO
                                        WHERE
                                            CNP.FIN_DEVOLUCAO.EMPRESA = B.EMPRESA
                                            AND CNP.FIN_DEVOLUCAO.REVENDA = B.REVENDA
                                            AND CNP.FIN_DEVOLUCAO.TITULO = B.TITULO
                                            AND CNP.FIN_DEVOLUCAO.DUPLICATA = B.DUPLICATA
                                            AND CNP.FIN_DEVOLUCAO.CLIENTE = B.CLIENTE
                                            AND CNP.FIN_DEVOLUCAO.TIPO = B.TIPO
                                    ),
                                    0
                                )
                            )
                        )
                    END SALDO,
                    (
                        COALESCE(
                            (
                                SELECT
                                    MAX(CNP.FIN_TITULO_PAGAMENTO.DTA_PAGAMENTO)
                                FROM
                                    CNP.FIN_TITULO_PAGAMENTO
                                WHERE
                                    CNP.FIN_TITULO_PAGAMENTO.EMPRESA = B.EMPRESA
                                    AND CNP.FIN_TITULO_PAGAMENTO.REVENDA = B.REVENDA
                                    AND CNP.FIN_TITULO_PAGAMENTO.TITULO = B.TITULO
                                    AND CNP.FIN_TITULO_PAGAMENTO.DUPLICATA = B.DUPLICATA
                                    AND CNP.FIN_TITULO_PAGAMENTO.CLIENTE = B.CLIENTE
                                    AND CNP.FIN_TITULO_PAGAMENTO.TIPO = B.TIPO
                                    AND B.STATUS <> 'DE'
                                    AND B.STATUS <> 'CA'
                                    AND B.STATUS <> 'EM'
                                    AND CNP.FIN_TITULO_PAGAMENTO.NRO_PAGAMENTO = (
                                        SELECT
                                            MAX(CNP.FIN_TITULO_PAGAMENTO.NRO_PAGAMENTO)
                                        FROM
                                            CNP.FIN_TITULO_PAGAMENTO
                                        WHERE
                                            CNP.FIN_TITULO_PAGAMENTO.EMPRESA = B.EMPRESA
                                            AND CNP.FIN_TITULO_PAGAMENTO.REVENDA = B.REVENDA
                                            AND CNP.FIN_TITULO_PAGAMENTO.TITULO = B.TITULO
                                            AND CNP.FIN_TITULO_PAGAMENTO.DUPLICATA = B.DUPLICATA
                                            AND CNP.FIN_TITULO_PAGAMENTO.CLIENTE = B.CLIENTE
                                            AND CNP.FIN_TITULO_PAGAMENTO.TIPO = B.TIPO
                                    )
                            ),
                            NULL
                        )
                    ) DTA_PAGAMENTO,
                    (
                	SELECT
                		G.DES_GRUPO_CLIENTE
                	FROM
                		CNP.FAT_GRUPO_CLIENTE G
                	WHERE
                		G.GRUPO_CLIENTE = A.GRUPO_CLIENTE
                    ) DES_GRUPO_CLIENTE,
                	SUBSTR(B.HISTORICO, 1, 200) HISTORICO,
                	(
                	SELECT
                		MAX(CPR.DES_PROVIDENCIA)
                	FROM
                		CNP.CAC_CONTATO CCO, 
                		CNP.CAC_PROVIDENCIA CPR,
                		CNP.GER_USUARIO GUS
                		WHERE
                		CCO.EMPRESA = B.EMPRESA
                		AND CCO.REVENDA = B.REVENDA
                		AND CCO.TITULO = B.TITULO
                		AND CCO.DUPLICATA = B.DUPLICATA
                		AND CCO.CLIENTE = B.CLIENTE
                		AND CCO.TIPO_DUPLICATA = B.TIPO		
                		AND CPR.EMPRESA = CCO.EMPRESA
                		AND CPR.REVENDA = CCO.REVENDA
                		AND CPR.CONTATO = CCO.CONTATO
                		AND CPR.PROVIDENCIA = (
                		SELECT
                			MAX(CPR1.PROVIDENCIA)
                		FROM
                			CNP.CAC_PROVIDENCIA CPR1
                		WHERE
                			CPR1.EMPRESA = CPR.EMPRESA
                			AND CPR1.REVENDA = CPR.REVENDA
                			AND CPR1.CONTATO = CPR.CONTATO
                         ) 
                         AND GUS.USUARIO = CPR.USUARIO
                	) INFORMACOES,
                	B.TIPO_ITEM
                FROM
                	CNP.FAT_CLIENTE A
                INNER JOIN (
                	SELECT
                		CNP.FIN_TITULO.EMPRESA,
                		CNP.FIN_TITULO.REVENDA,
                		CNP.FIN_TITULO.TITULO,
                		CNP.FIN_TITULO.DUPLICATA,
                		CNP.FIN_TITULO.CLIENTE,
                		CNP.FIN_TITULO.OPERACAO,
                		CNP.FIN_TITULO.TIPO,
                		CNP.FIN_TITULO.BANCO,
                		CNP.FIN_TITULO.HISTORICO,
                		CNP.FIN_TITULO.VAL_TITULO,
                		CNP.FIN_TITULO.PAGAR_RECEBER,
                		CNP.FIN_TITULO.DTA_EMISSAO,
                		CNP.FIN_TITULO.DTA_VENCIMENTO,
                		CNP.FIN_TITULO.STATUS,
                		CNP.FIN_TITULO.DEPARTAMENTO,
                		CNP.FIN_TITULO.ORIGEM,
                		CNP.FIN_TITULO.PLANO_CONTRATO,
                		CNP.FIN_TITULO.TIPO_ITEM,
                		CNP.FIN_TITULO.USUARIO,
                		CNP.FIN_TITULO.PROPOSTA,
                		CNP.FIN_TITULO.LOCALIZACAO
                	FROM
                		CNP.FIN_TITULO
                 ) B ON
                	(B.CLIENTE = A.CLIENTE)
                LEFT JOIN CNP.FAT_MOVIMENTO_CAPA FMC ON
                	(
                        FMC.EMPRESA = B.EMPRESA
                		AND FMC.REVENDA = B.REVENDA
                		AND FMC.OPERACAO = B.OPERACAO
                 )
                LEFT JOIN CNP.VEI_PROPOSTA VP ON
                	(
                        VP.EMPRESA = FMC.EMPRESA
                		AND VP.REVENDA = FMC.REVENDA
                		AND VP.CONTATO = FMC.CONTATO
                    )
                LEFT JOIN CNP.VEI_AVALIACAO_PROPOSTA AVP ON
                	(
                        AVP.EMPRESA = B.EMPRESA
                		AND AVP.REVENDA = B.REVENDA
                		AND AVP.PROPOSTA = VP.PROPOSTA
                    )
                LEFT JOIN CNP.VEI_AVALIACAO AV ON
                	(
                        AV.EMPRESA = AVP.EMPRESA
                		AND AVP.AVALIACAO = AV.AVALIACAO
                    )
                LEFT JOIN CNP.FAT_PESSOA_FISICA PF ON
                	PF.CLIENTE = A.CLIENTE
                LEFT JOIN CNP.FAT_PESSOA_JURIDICA PJ ON
                	PJ.CLIENTE = A.CLIENTE
                WHERE
                    B.TIPO IN ('CP')
                	AND B.STATUS IN ('EM', 'PP')
                	--AND B.ORIGEM IN (10334, 80160)
                ORDER BY
                	A.NOME,
                	B.TITULO)


        SELECT 
        (dw.EMPRESA ||'.'|| dw.REVENDA) loja,
        dw.NOME,

        (CASE
        WHEN dw.ORIGEM IN (20128, 20130, 20134, 20136, 20146, 20151, 20155,20157, 20666, 20668) THEN 'Empréstimo'
        WHEN dw.ORIGEM IN (20261,20262,20264,20282,20283,20284,20285,20286,20287,20288,20289, 20290,20291,20292,20293,20294,20297,20298,20299,20300,20301) THEN  'Imposto' 
        WHEN dw.ORIGEM IN (20005) THEN 'Clientes'
        WHEN dw.ORIGEM IN (20004,20006,20009,20660,80045,90077,90091,90092) THEN 'Fábrica'
        WHEN dw.ORIGEM IN (20202, 20219) THEN 'Fornecedor'
        WHEN dw.ORIGEM IN (90006,90007,90008,90029,90031,90032,90033,90034,90037,90038,90260) THEN 'Vendas'
        ELSE 'Outros'
        END) categoria,

        TO_CHAR(dw.DTA_EMISSAO, 'DD/MM/YYYY') EMISSAO,
        TO_CHAR(dw.DTA_VENCIMENTO,'DD/MM/YYYY') VENCIMENTO,
        (
        SELECT 
        gd.NOME
        FROM CNP.GER_DEPARTAMENTO gd 
        WHERE 
        gd.EMPRESA=dw.EMPRESA AND 
        gd.REVENDA=dw.REVENDA AND
        gd.DEPARTAMENTO=dw.DEPARTAMENTO
        ) DEPARTAMENTO,
        COALESCE(dw.INFORMACOES,'-') INFORMACAO,
        dw.VAL_TITULO,
        dw.SALDO


        FROM dw 

        WHERE trunc(dw.DTA_VENCIMENTO) = trunc(current_date)
        '''
        with self._database.session_factory() as session:
            data = pd.read_sql(sql_query, session.connection())
            data_transformed = self.__transform_data(data, 'pdf')
            self.__transform_data_in_pdf_file(data_transformed)
            return data_transformed
